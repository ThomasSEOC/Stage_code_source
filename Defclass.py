#Definition des classes :
from abc import *
from enum import Enum
from strenum import StrEnum
from openpyxl import *
from openpyxl.utils import *
from mailmerge import MailMerge

# Commentaires + test unitaire + google colab


class ExcelFile :
    """Define functions according to the API used (here openpyxl) to be
    independant from it in the rest of the code.
    
    """
    
    def __init__(self,name):
        self.name = name
        self.wb = load_workbook(self.name,data_only=True)
        
    def set_current_sheet(self,sheet_name):
        """Open the targeted sheet"""
        self.sheet = self.wb[sheet_name]
        
    def get_max_row(self) :
        """Return the number of rows that have been modified in the sheet"""
        return(self.sheet.max_row)
    
    def get_max_column(self) :
        """Return the number of columns that have been modified in the sheet"""
        return(self.sheet.max_column)
               
    def _convert_type(self,cell) :
        """Treat the data type of a cell
        If the data type is not numeric/string/date, returns the data type discovered
        """
        if (cell.data_type == 'n') :
            return(ValueType.NUMERIC)
            
        if (cell.data_type == 's'):
            return(ValueType.STRING)
        
        if (cell.data_type == 'd'):
            return(ValueType.DATE)
        
        else :
            raise ValueError(f'{cell.data_type} non défini')
    
    def get_cell_value(self,i,j):
        """Return the value on row 'i', column 'j' """
        return self.sheet.cell(i,j).value
    
    def get_cell_type(self,i,j):
        """Return the data type of the value on row 'i', column 'j' """
        return self._convert_type(self.sheet.cell(i,j))
        
    def _get_named_cell(self,cell_name):
        """Return the adress of the cell getting the name 'cell_name' """
        info_cell=self.wb.defined_names[cell_name]
        cell_untreated = info_cell.attr_text
        sh = cell_untreated.split('!')[0]
        location = cell_untreated.split('!')[1]
        var =self.wb[sh][location]
        return(var)
           
    def get_cell_named_value(self,cell_name):
        """Return the value at the adress given by 'get_named_cell' """
        var = self._get_named_cell(cell_name)
        return(var.value)
    
    def get_cell_named_type(self,cell_name):
        """Return the data type of the value at the adress given by 'get_named_cell' """
        var = self._get_named_cell(cell_name)
        return self._convert_type(var)

class Variable :
    
    def __init__(self,var_name,value=None):
        self.var_name=var_name
        self.value=value

class AbstractDatasource(ABC) :

    @abstractmethod
    def get_var(self,var_name) -> Variable :
        pass
    
class AbstractMultiDatasource(AbstractDatasource):
    
    @abstractmethod
    def next_item(self):
        pass

class ValueFormat(StrEnum):
    DATE = 'Date'
    TEXTE = 'Texte'
    NOMBRE = 'Nombre'
    MONETAIRE = 'Monétaire'
    POURCENTAGE = 'Pourcentage'
    CODE_POSTAL = 'Code postal'

class ValueType(Enum):  
    NUMERIC = 1     
    STRING = 2  
    FORMULA = 3
    DATE = 4

class Value :
    
    def __init__(self,value_type,value):
        self.value = value
        self.value_type = value_type

class Formula(Variable) :
    
    def __init__(self,var_name,value,value_format):
        self.var_name=var_name
        self.value=Value(ValueType.FORMULA,value)
        self.value_format = value_format

    def get_var_names(self):
        separators = "()/*-+"
        ch = self.value.value
        for k in separators :
            ch = ch.replace(k,' ')
        return list(filter(None,ch.split(' ')))

#a = Formula('VarTest','(Fiche_Client.Dépla+(Fiche_Client.totalrestauHT*Fiche_Client.coutformHT))/Fiche_Client.Txtva','Monétaire').get_var_names()

class Formater :
    
    def __init__(self,value,format_value):
        self.value = value
        self.format_value = format_value
        
    def _verif_format_date(self):
        ch = self.value
        if (ch[4]=='-' and ch[7]=='-' and ch[10]==' ' and ch[13]==':' and ch[16]==':'):
            return(True)
        else :
            return(False)
    
    def _verif_format_cp(self):
        if (len(str(self.value)) == 5):
            return(True)
        else :
            return(False)
    
    def formating(self):
        
        if (self.format_value == ValueFormat.DATE) :
            if (self._verif_format_date) :
                L = self.value.split(' ')[0].split('-')
                displayed_value = L[2]+'/'+L[1]+'/'+L[0]
                return(displayed_value)
            else : 
                raise ValueError(f"la valeur {self.value} ne convient pas au format {self.format_value}")
        
        if (self.format_value == ValueFormat.TEXTE) :
            return(str(self.value))
        
        if (self.format_value == ValueFormat.NOMBRE) :
            return(str(self.value))
        
        if (self.format_value == ValueFormat.MONETAIRE) :
            return(str(self.value) + ' €')
        
        if (self.format_value == ValueFormat.POURCENTAGE) :
            return(str(self.value) + ' %')
        
        if (self.format_value == ValueFormat.CODE_POSTAL) :
            if (self._verif_format_cp()):
                cp = str(self.value)
                return(cp[0]+cp[1]+' '+cp[2]+cp[3]+cp[4])
            else :
                raise ValueError(f"la valeur {self.value} ne correspond pas au format {self.format_value}")
    
    
class LimesurveyExcelDatasource(AbstractDatasource):
    
    def __init__(self,name,path,sh='Feuil1'):
        self.name=name
        self.path=path
        self.excel = ExcelFile(path)
        self.excel.set_current_sheet(sh)
    
    def get_var(self,var_name) -> Variable :
        col = 1
        while (self.excel.get_cell_value(1,col) != var_name):
            if (self.excel.get_cell_value(1,col) == None):
                raise ValueError(f"la variable {var_name} de la datasource {self.name} n'existe pas ou est mal orthographiée")
            col +=1
        value = Value(self.excel.get_cell_type(2, col),self.excel.get_cell_value(2,col))
        return Variable(var_name,value)


class LimesurveyMultiExcelDatasource(AbstractMultiDatasource) :
    
    def __init__(self,code_client,name,path,sh='Feuil1') :
        self.code_client = code_client
        self.name = name
        self.path = path
        self.excel = ExcelFile(path)
        self.excel.set_current_sheet(sh)
        self.line = 2
        
    def _seek_client(self,code_client):
         while (self.excel.get_cell_value(self.line, 1) != self.code_client) :
            if (self.excel.get_cell_value(self.line,1) == None):
                raise ValueError(f"le code client {code_client} de la datasource {self.name} n'existe pas ou  est mal orthographiée")
            self.line += 1
        
    def get_var(self,var_name) :
        column = 1

        #Recherche du bon client
        if (self.line == 2):
            self._seek_client(self.code_client)
            
        #Recherche de la colonne "var_name"    
        while (self.excel.get_cell_value(1,column) != var_name):
            if (self.excel.get_cell_value(1,column) == None):
                raise ValueError(f"la variable {var_name} de la datasource {self.name} n'existe pas ou est mal orthographiée")
            column +=1
            
        value = Value(self.excel.get_cell_type(self.line,column),self.excel.get_cell_value(self.line,column))
        
        return(Variable(var_name,value))
    
    def next_item(self):
         self.line += 1
         if (self.excel.get_cell_value(self.line, 1) != self.code_client) :
             return None 
         return(self)
    

class InfoClientDatasource(AbstractDatasource) :
    
    def __init__(self,name,path,sh='InfoClient') :

        self.name = name
        self.path = path
        self.excel = ExcelFile(path)
        self.excel.set_current_sheet(sh)
        
    def get_var(self,named_value) -> Variable :
        try :
            value = Value(self.excel.get_cell_named_type(named_value),self.excel.get_cell_named_value(named_value))
        except KeyError :
            raise ValueError(f"La zone nommée {named_value} n'existe pas dans la datasource {self.name} ou est mal orthographiée")
        return(Variable(named_value,value))

class DatasourceType(StrEnum):
    CHAMP = "Champ"
    CHAMPMULTI = "Champmulti"
    NOMME = "Nommé"
        
 
class DatasourceManager :
    
    def __init__(self,path,sheet,code_client = None ): # Définir tous les onglets source par 'Source'
        self.path = path
        self.excel = ExcelFile(path)
        self.excel.set_current_sheet(sheet)
        self.sources = {}
        self.code_client = code_client
        self.multi_name = None
        self._read_spec()
        if (self.multi_name != None and self.code_client == None):
            raise ValueError('Il manque la propriété "code client')
            
        
    def _read_spec(self):
        row = 2
        while (self.excel.get_cell_value(row,1) != None) :
            name = self.excel.get_cell_value(row,1)
            file = self.excel.get_cell_value(row,2)
            tp = self.excel.get_cell_value(row,3)
            sheet = self.excel.get_cell_value(row,4)
            if (tp == DatasourceType.CHAMP):
                self.sources[name] = LimesurveyExcelDatasource(name,file,sheet)
            elif (tp == DatasourceType.CHAMPMULTI):
                self.multi_name = name
                #print(self.multi_name)
                self.sources[name] = LimesurveyMultiExcelDatasource(self.code_client,name,file,sheet)
            elif (tp == DatasourceType.NOMME):
                self.sources[name] = InfoClientDatasource(name,file,sheet)
            else :
                raise ValueError(f"Le type de datasource {tp} n'est pas connu")
            row +=1
        
    def get_var(self,src_name,var_name):
        if src_name in self.sources :
            return self.sources[src_name].get_var(var_name)
        else :
            raise ValueError(f"La datasource {src_name} n'existe pas")
    
    def next_item(self):
        #print(self.multi_name)
        if (self.multi_name != None ) :
            return self.sources[self.multi_name].next_item()
        else :
            return None


class FormulaManager :
    
    def __init__(self,path,sheet_calc,datasource_manager): # Définir tous les onglets source par 'Source'
        self.path = path
        self.datasource_manager = datasource_manager
        self.excel = ExcelFile(path)
        self.sheet_calc = sheet_calc
        self.formulas = []
        self.excel.set_current_sheet(self.sheet_calc)
        self._read_spec()
        self.vars_dictionary = {}
        
    def _read_spec(self):
        
        row = 2
        while (self.excel.get_cell_value(row,1) != None) :
            var_name = self.excel.get_cell_value(row,1)
            var_formula = self.excel.get_cell_value(row,2)
            format_value = self.excel.get_cell_value(row,3)
            formula = Formula(var_name,var_formula,format_value)
            self.formulas.append(formula)
            row += 1
            
    
    def next_vars_dictionary(self) -> dict :
        
        if (self.vars_dictionary) :
            r = self.datasource_manager.next_item()
            if (r == None):
                return(None)
            
            
        self.vars_dictionary = {}
        for formula in self.formulas :
            formula_exp = formula.value.value
            var_names = formula.get_var_names()
            for var_name in var_names :
                if '.' in var_name :
                    src_name = var_name.split('.')[0]
                    src_var_name = var_name.split('.')[1]
                    var = self.datasource_manager.get_var(src_name,src_var_name)
                    formula_exp = formula_exp.replace(var_name,str(var.value.value or ''))
                else :
                    formula_exp = formula_exp.replace(var_name,str(self.vars_dictionary[var_name]))
            if(len(var_names) > 1) :
                self.vars_dictionary[formula.var_name] = eval(formula_exp)
            else :
                self.vars_dictionary[formula.var_name] = formula_exp
            
        for formula in self.formulas :
            self.vars_dictionary[formula.var_name] = Formater(self.vars_dictionary[formula.var_name],formula.value_format).formating()
            
        return(self.vars_dictionary)
                    
                    

class Composer :
    
    def __init__(self,word_path,formula_manager,doc_gen_path):
        self.word_path = word_path
        self.formula_manager = formula_manager
        self.doc_gen_path = doc_gen_path
    
    def _compose_doc(self):
        n = 0
        while(True):
            fm_next_var = self.formula_manager.next_vars_dictionary()
            if(fm_next_var == None):
                print(f"Les documents ont été créés à l'emplacement cible")
                break
            n += 1
            template = self.word_path
            doc = MailMerge(template)
            doc.merge(**fm_next_var)
            
            doc_gen_path = self.doc_gen_path.replace('.docx',str(n)+'.docx')            
            doc.write(self.doc_gen_path)
            if(n > 10):
                break
            
    
    def compose(self):
        self._compose_doc()
        

class Generator :
    
    def __init__(self,word_path,target_path,doc_gen_path,code_client):
        self.word_path = word_path
        self.target_path = target_path
        self.code_client = code_client
        self.doc_gen_path = doc_gen_path
        
    def doc_gen(self) :
        fm = FormulaManager(self.target_path,'Calcul',DatasourceManager(self.target_path, 'Source',self.code_client))
        word_path = "D:\\Users_D\\docs_bettersupply\\CONVENTION_publi\\Convention_ThomasV1_publi.docx"
        target_path = "D:\\Users_D\\docs_bettersupply\\CONVENTION_publi\\Test_CONVENTION.docx"
        Composer(self.word_path,fm,self.doc_gen_path).compose()

# word_path = "D:\\Users_D\\docs_bettersupply\\CONVENTION_publi\\Convention_ThomasV1_publi.docx"
# doc_gen_path = "D:\\Users_D\\docs_bettersupply\\CONVENTION_publi\\Test_CONVENTION.docx"
# target_path = "D:\\Users_D\\docs_bettersupply\\CONVENTION_publi\\CONVENTION_publi.xlsx"
# "HRMI01"
        
        
def main() :
    a = FormulaManager('CONVENTION_publi.xlsx','Calcul',DatasourceManager('CONVENTION_publi.xlsx', 'Source','HRMI01'))
    word_path = "D:\\Users_D\\docs_bettersupply\\CONVENTION_publi\\Convention_ThomasV1_publi.docx"
    target_path = "D:\\Users_D\\docs_bettersupply\\CONVENTION_publi\\Test_CONVENTION.docx"
    Composer(word_path,a,target_path).compose()
    
def main2() :
    a = FormulaManager('CONVENTION_publi.xlsx','Calcul',DatasourceManager('CONVENTION_publi.xlsx', 'Source','BETT01'))
    word_path = "D:\\Users_D\\docs_bettersupply\\CONVENTION_publi\\Convention_ThomasV1_publi.docx"
    target_path = "D:\\Users_D\\docs_bettersupply\\CONVENTION_publi\\Test_BETT.docx"
    Composer(word_path,a,target_path).compose()
    

def main3() :
    a= FormulaManager('CDC_publi.xlsx','Calcul',DatasourceManager('CDC_publi.xlsx','Source'))
    word_path = 'CDC_publi.docx'
    target_path = "D:\\Users_D\\docs_bettersupply\\CONVENTION_publi\\Test_CDC.docx"
    Composer(word_path,a,target_path).compose()

def test_multi():
    ds = DatasourceManager('CONVENTION_publi.xlsx', 'Source','HRMI01')
    fm = FormulaManager('CONVENTION_publi.xlsx','Calcul',ds)
    fm.next_vars_dictionary()
    fm.next_vars_dictionary()
    fm.next_vars_dictionary()
    print('multi : ',fm.next_vars_dictionary())
    
    ds = DatasourceManager('CDC_publi.xlsx', 'Source','HRMI01')
    fm = FormulaManager('CDC_publi.xlsx','Calcul',ds)
    fm.next_vars_dictionary()
    print('simple : ',fm.next_vars_dictionary())
    
    
    
    
    